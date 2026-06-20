#!/usr/bin/env python3
"""Smoke B5: nhận document (không-ảnh) → ép claude-path + bake hướng dẫn đọc.
Import bridge THẬT, monkeypatch đầu nối Telegram/Claude → KHÔNG gọi mạng, KHÔNG tốn token, KHÔNG restart."""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
import lucy_bridge as lb

PASS = []
def chk(name, cond):
    PASS.append(cond)
    print(("✅" if cond else "❌"), name)

# ── files mẫu thật để chứng minh đọc được ──
import openpyxl, docx
with open("/tmp/lucy-smoke.csv", "w") as f:
    f.write("thang,doanh_thu\n6/2026,1250000\n")
wb = openpyxl.Workbook(); wb.active.append(["Thang","Doanh thu"]); wb.active.append(["6/2026",1250000]); wb.save("/tmp/lucy-smoke.xlsx")
d = docx.Document(); d.add_paragraph("Bao cao FitCity thang 6"); d.save("/tmp/lucy-smoke.docx")

# ── monkeypatch đầu nối ngoài ──
cap = {}
lb.ALLOWED = ""                                   # bỏ khóa user cho smoke
lb.tg_download = lambda fid, prefix="doc": "/tmp/lucy-smoke.csv"
lb.send = lambda cid, t: None
lb.send_id = lambda cid, t: 1
lb.edit = lambda *a, **k: None
lb.reply = lambda cid, t: cap.update(reply=t)
lb.recall_prefetch = lambda t: ""
lb.episodic_log = lambda *a, **k: None
lb._token_guard_hard = lambda: False
lb.resolve_persona_text = lambda p: ""
lb._save = lambda s: None
lb._take_claude_seed = lambda c: ""
lb.claude_hist_append = lambda *a, **k: None
lb.should_compress = lambda c: False
lb._load_prefs = lambda: {"123": {"model": "deepseek:r1"}}   # ép pref = LANE → test force-claude
def fake_stream(prompt, sid, model, persona, on_delta):
    cap["prompt"] = prompt; cap["model"] = model
    return ("sid-smoke", "Em đã đọc file rồi ạ.", None)
lb.run_claude_stream = fake_stream
# lane KHÔNG được gọi với document → nếu bị gọi là FAIL
lb.run_lane = lambda *a, **k: (_ for _ in ()).throw(AssertionError("document KHÔNG được rớt xuống lane!"))

# ── CASE 1: document không-ảnh (.csv) + caption ──
msg = {"chat": {"id": 123}, "from": {"id": "999"},
       "document": {"file_id": "abc", "file_name": "baocao.csv", "mime_type": "text/csv"},
       "caption": "tóm tắt giúp em"}
cap.clear()
lb.handle(msg, {})
p = cap.get("prompt", "")
chk("CASE1 gọi claude-path (không rớt lane)", "prompt" in cap)
chk("CASE1 model bị ÉP về sonnet (dù pref=deepseek:r1)", cap.get("model") == "sonnet")
chk("CASE1 prompt chứa path file đã tải", "/tmp/lucy-smoke.csv" in p)
chk("CASE1 prompt giữ caption chủ nhân", "tóm tắt giúp em" in p)
chk("CASE1 prompt có hướng dẫn xlsx (openpyxl)", "openpyxl" in p)
chk("CASE1 prompt có hướng dẫn docx", "docx" in p)
chk("CASE1 prompt có hướng dẫn ảnh nhúng", "ẢNH nhúng" in p)
chk("CASE1 prompt có tên file gốc", "baocao.csv" in p)

# ── CASE 2: regression — ảnh (document image/*) vẫn đi nhánh vision, KHÔNG đi nhánh doc ──
lb.tg_download = lambda fid, prefix="img": "/tmp/lucy-smoke.png"
msg2 = {"chat": {"id": 123}, "from": {"id": "999"},
        "document": {"file_id": "img1", "file_name": "anh.png", "mime_type": "image/png"}}
cap.clear()
lb.handle(msg2, {})
p2 = cap.get("prompt", "")
chk("CASE2 ảnh vẫn vào nhánh vision (1 ẢNH)", "1 ẢNH" in p2 and "/tmp/lucy-smoke.png" in p2)

# ── CASE 3: đọc nội dung thật từng loại (chứng minh lib chạy, không nói suông) ──
import subprocess
csv_txt = open("/tmp/lucy-smoke.csv").read()
chk("CASE3 csv đọc được", "1250000" in csv_txt)
wb2 = openpyxl.load_workbook("/tmp/lucy-smoke.xlsx")
chk("CASE3 xlsx đọc được", 1250000 in [c for r in wb2.active.iter_rows(values_only=True) for c in r])
d2 = docx.Document("/tmp/lucy-smoke.docx")
chk("CASE3 docx đọc được", any("FitCity" in p.text for p in d2.paragraphs))

print(f"\n{'='*40}\nKẾT QUẢ: {sum(PASS)}/{len(PASS)} PASS")
sys.exit(0 if all(PASS) else 1)
